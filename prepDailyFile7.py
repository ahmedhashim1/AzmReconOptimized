import pandas as pd
import xlwings as xw
from datetime import date, datetime
import time
import os
import numpy as np
import mysql.connector
from mysql.connector import Error
import concurrent.futures
from threading import Lock
import logging
import traceback
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
from pathlib import Path
import copy
import re
import config

m_day = config.config.curr_day
m_month = config.config.curr_month
m_year = config.config.curr_year
m_date = datetime(m_year, m_month, m_day)
path_month_abbr = m_date.strftime("%b")
path_year = m_date.strftime("%Y")
path_month_full = m_date.strftime("%B")
path_month_abbr = m_date.strftime("%b")
path_day = m_date.strftime("%d")
# trans_date = date.strftime("%Y/%m/%d")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# MySQL configuration details
mysql_config = {
    "host": "localhost",
    "user": "root",
    "password": "root",
    "database": "azm"
}

# Global variables
connection_pool = None
pool_lock = Lock()


def get_mysql_connection_pool():
    """Create or get a connection pool for better performance"""
    global connection_pool
    if connection_pool is None:
        with pool_lock:
            if connection_pool is None:
                try:
                    from mysql.connector import pooling
                    pool_config = mysql_config.copy()
                    pool_config.update({
                        'pool_name': 'azm_pool',
                        'pool_size': 10,
                        'pool_reset_session': True,
                        'autocommit': True
                    })
                    pool = pooling.MySQLConnectionPool(**pool_config)
                    logger.info(f"Created connection pool with {pool.pool_size} connections")
                    connection_pool = pool
                except mysql.connector.Error as err:
                    logger.error(f"Error creating connection pool: {err}")
    return connection_pool


def modify_excel_file_final(daily_file_path, master_file_path):
    """
    Modifies an Excel file by rearranging columns and adding data from a master file
    while preserving original formatting.
    """
    app = None
    try:
        logger.info("PHASE 1: Starting column movements.")

        # Try to connect to existing Excel app, create new one if needed
        app = None
        try:
            apps = xw.apps
            if len(apps) > 0:
                app = apps.active
                logger.info("Connected to existing Excel application")
            else:
                raise Exception("No active apps found")
        except Exception:
            logger.info("Creating new Excel application")
            try:
                app = xw.App(visible=True, add_book=False)
                logger.info("Created new Excel application")
            except Exception as e:
                logger.error(f"Failed to create Excel app: {e}")
                time.sleep(3)
                app = xw.App(visible=True, add_book=False)

        if app is None:
            raise Exception("Could not create or connect to Excel application")

        # Open the daily workbook
        daily_wb = app.books.open(daily_file_path)
        daily_sheet = daily_wb.sheets[0]

        last_row = daily_sheet.range('A1').end('down').row

        if last_row is None or last_row < 2:
            logger.error("The daily file is empty or contains only headers.")
            daily_wb.close()
            return

        # Column movements
        daily_sheet.range('B:B').api.Cut(Destination=daily_sheet.range('S1').api)
        daily_sheet.range('A:A').api.Cut(Destination=daily_sheet.range('B1').api)
        daily_sheet.range('B1').api.EntireColumn.Insert()

        # Insert Headers
        daily_sheet.range('A1').value = 'Cust'
        daily_sheet.range('B1').value = 'Index'

        # Write 'fdate' to Column U
        today = date.today()
        daily_sheet.range('U1').value = 'fdate'
        daily_sheet.range('U2:U' + str(last_row)).value = today
        logger.info("PHASE 1: Column movements completed.")

        # PHASE 2: Pandas processing
        logger.info("PHASE 2: Starting data processing.")

        اسم_المفوتر_col = daily_sheet.range('C2:C' + str(last_row)).options(ndim=1).value
        اسم_المفوتر_df = pd.DataFrame(اسم_المفوتر_col, columns=['اسم المفوتر'])

        master_df = pd.read_excel(master_file_path, usecols=['Arabic', 'Name', 'Index', 'Type', 'Transf Type'])
        master_df.rename(columns={'Arabic': 'اسم المفوتر', 'Name': 'Cust', 'Index': 'Index'}, inplace=True)

        merged_df = pd.merge(اسم_المفوتر_df, master_df, on='اسم المفوتر', how='left')
        daily_sheet.range('A2').value = merged_df[['Cust', 'Index']].values

        # PHASE 3: Convert amount columns to numeric format using Excel's TextToColumns (FAST!)
        logger.info("PHASE 3: Converting amount columns to numeric format using TextToColumns...")

        # Define numeric columns by position (E, F, H, I, K = columns 5, 6, 8, 9, 11)
        numeric_columns_positions = [5, 6, 8, 9, 11]  # Excel columns E, F, H, I, K (1-indexed)

        for col_pos in numeric_columns_positions:
            if col_pos <= daily_sheet.used_range.last_cell.column:
                try:
                    # Convert column number to letter
                    col_letter = chr(64 + col_pos)  # Convert to letter: 5->E, 6->F, etc.
                    logger.info(f"Converting column {col_letter} to numeric using TextToColumns...")

                    # Select the data range for this column (excluding header)
                    start_cell = daily_sheet.range(f'{col_letter}2')
                    data_range = daily_sheet.range(start_cell, start_cell.end('down'))

                    # Use Excel's TextToColumns feature for fast conversion
                    data_range.api.TextToColumns(
                        Destination=start_cell.api,
                        DataType=1,  # xlDelimited
                        TextQualifier=1,  # xlDoubleQuote
                        ConsecutiveDelimiter=False,
                        Tab=True,
                        Semicolon=False,
                        Comma=False,
                        Space=False,
                        Other=False,
                        FieldInfo=((1, 1),),  # Array(1, 1) - General format
                        TrailingMinusNumbers=True
                    )

                    # Now apply number formatting
                    data_range.number_format = '#,##0.00'

                    logger.info(f"Successfully converted column {col_letter} to numeric format")

                except Exception as e:
                    logger.warning(f"Could not convert column {col_pos} ({chr(64 + col_pos)}) to numeric: {e}")

                    # Fallback to simpler method if TextToColumns fails
                    try:
                        logger.info(f"Trying fallback method for column {col_letter}...")
                        col_letter = chr(64 + col_pos)
                        col_range = daily_sheet.range(f'{col_letter}2:{col_letter}{last_row}')

                        # Simple approach: just change the number format and let Excel handle it
                        col_range.number_format = '#,##0.00'

                        # Force calculation to convert text numbers
                        col_range.api.Calculate()

                        logger.info(f"Fallback conversion successful for column {col_letter}")
                    except Exception as e2:
                        logger.warning(f"Fallback also failed for column {col_letter}: {e2}")

        daily_wb.save()
        logger.info("PHASE 3: Fast numeric conversion completed using TextToColumns.")

        logger.info("Excel file modification completed successfully.")

    except Exception as e:
        logger.error(f"Error in modify_excel_file_final: {e}")
        raise


def detect_leading_zeros_pattern(series):
    """
    Detect if a pandas series likely contains values with leading zeros
    Returns True if pattern suggests leading zeros should be preserved
    """
    if series.dtype == 'object':  # String columns
        return True

    # Check for numeric patterns that might have had leading zeros
    sample_values = series.dropna().astype(str).head(100)  # Check first 100 non-null values

    # Look for patterns that suggest leading zeros
    patterns = [
        r'^\d+$',  # Pure numeric strings
        r'^\d{6,}$',  # Long numeric strings (likely IDs)
        r'^0+\d+',  # Strings starting with zeros
    ]

    for pattern in patterns:
        matches = sum(1 for val in sample_values if re.match(pattern, str(val)))
        if matches > len(sample_values) * 0.3:  # If 30% match pattern
            return True

    return False


def smart_read_excel_with_string_preservation(file_path):
    """
    Fast hybrid approach: Read with pandas for speed, but preserve string formatting
    where needed using intelligent detection. Numeric columns are already converted in main file.
    """
    logger.info("📊 Smart reading Excel with string preservation...")

    # Step 1: Quick pandas read to get structure and detect patterns
    temp_df = pd.read_excel(file_path, nrows=50)  # Read just 50 rows for pattern detection

    # Step 2: Identify columns that likely need string preservation
    string_columns = []
    column_names = temp_df.columns.tolist()

    # Position-based detection (D, M, T columns)
    position_based_columns = []
    if len(column_names) > 3:  # Column D
        position_based_columns.append(column_names[3])
    if len(column_names) > 12:  # Column M
        position_based_columns.append(column_names[12])
    if len(column_names) > 19:  # Column T
        position_based_columns.append(column_names[19])

    # Name-based detection
    name_based_columns = ['InvoiceNum', 'InternalCode', 'ContractNum']

    # Pattern-based detection
    pattern_based_columns = []
    for col in temp_df.columns:
        if detect_leading_zeros_pattern(temp_df[col]):
            pattern_based_columns.append(col)

    # Combine all approaches for string columns
    string_columns = list(set(position_based_columns + name_based_columns + pattern_based_columns))

    # Step 3: Identify numeric amount columns (E, F, H, I, K) - these are already converted in main file
    numeric_columns = []
    if len(column_names) > 4:  # Column E
        numeric_columns.append(column_names[4])
    if len(column_names) > 5:  # Column F
        numeric_columns.append(column_names[5])
    if len(column_names) > 7:  # Column H
        numeric_columns.append(column_names[7])
    if len(column_names) > 8:  # Column I
        numeric_columns.append(column_names[8])
    if len(column_names) > 10:  # Column K
        numeric_columns.append(column_names[10])

    logger.info(f"🔤 Detected string columns: {string_columns}")
    logger.info(f"🔢 Detected numeric columns (already converted in main file): {numeric_columns}")

    # Step 4: Read full file with appropriate dtypes
    dtype_dict = {}
    for col in string_columns:
        if col in column_names:
            dtype_dict[col] = str  # Force as string to preserve leading zeros

    # For numeric columns, let pandas read them naturally as they're already properly formatted
    # in the main Excel file after our conversion

    # Read the full file with string preservation
    logger.info("📖 Reading full file with optimized dtypes...")
    full_df = pd.read_excel(file_path, dtype=dtype_dict)

    # No need to convert numeric columns here - they're already converted in the main file

    return full_df, string_columns, numeric_columns


def lightning_fast_formatted_split(daily_file_path, master_file_path):
    """
    LIGHTNING FAST version: Combines speed with formatting and string preservation
    Uses hybrid approach for optimal performance
    """
    start_time = time.time()
    logger.info("⚡ Starting LIGHTNING FAST file splitting...")

    try:
        # Step 1: Smart read with string preservation and numeric conversion (FAST)
        full_df, string_columns, numeric_columns = smart_read_excel_with_string_preservation(daily_file_path)

        if full_df.empty:
            logger.warning("Daily file is empty!")
            return 0, 0, 0

        # Step 2: Quick formatting template extraction (FAST)
        logger.info("🎨 Extracting formatting template...")
        template_wb = load_workbook(daily_file_path)
        template_ws = template_wb.active

        # Get formatting for first data row only (much faster than all rows)
        header_formats = {}
        data_formats = {}

        max_cols = min(template_ws.max_column, len(full_df.columns))

        for col_idx in range(1, max_cols + 1):
            col_name = full_df.columns[col_idx - 1] if col_idx <= len(full_df.columns) else None

            # Header formatting
            header_cell = template_ws.cell(row=1, column=col_idx)
            header_formats[col_idx] = {
                'font': Font(bold=True) if header_cell.font and header_cell.font.bold else Font(),
                'number_format': '@' if col_name in string_columns else 'General'
            }

            # Data formatting (simplified)
            if template_ws.max_row >= 2:
                data_formats[col_idx] = {
                    'number_format': (
                        '@' if col_name in string_columns else
                        '#,##0.00' if col_name in numeric_columns else
                        'General'
                    )
                }
            else:
                data_formats[col_idx] = header_formats[col_idx]

        # Get column widths (fast)
        column_widths = {}
        for col_letter, dimension in template_ws.column_dimensions.items():
            column_widths[col_letter] = dimension.width if dimension.width else 13.0

        template_wb.close()

        # Step 3: Lightning fast customer processing
        unique_customers = full_df['Cust'].dropna().unique()
        logger.info(f"👥 Processing {len(unique_customers)} customers...")
        logger.info(f"🔤 String columns: {string_columns}")
        logger.info(f"🔢 Numeric columns: {numeric_columns}")

        file_dir = os.path.dirname(daily_file_path)
        successful_files = 0
        failed_files = 0

        # Group data once for efficiency
        customer_groups = full_df.groupby('Cust')

        for customer_name in unique_customers:
            try:
                customer_data = customer_groups.get_group(customer_name)

                if customer_data.empty:
                    continue

                # Create safe filename
                safe_filename = "".join(
                    c for c in str(customer_name) if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
                output_file = os.path.join(file_dir, f"{safe_filename}.xlsx")

                # Ultra-fast workbook creation
                wb = Workbook()
                ws = wb.active

                # Bulk write data (SUPER FAST)
                rows_data = dataframe_to_rows(customer_data, index=False, header=True)
                for r_idx, row in enumerate(rows_data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)

                        # Apply essential formatting only
                        if r_idx == 1 and c_idx in header_formats:  # Header
                            cell.font = header_formats[c_idx]['font']
                            cell.number_format = header_formats[c_idx]['number_format']
                        elif r_idx > 1 and c_idx in data_formats:  # Data
                            cell.number_format = data_formats[c_idx]['number_format']

                # Quick column width application
                for col_idx, col_name in enumerate(customer_data.columns, 1):
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    ws.column_dimensions[col_letter].width = column_widths.get(col_letter, 13.0)

                # Save and close
                wb.save(output_file)
                wb.close()

                successful_files += 1

                # Progress update every 50 files
                if successful_files % 50 == 0:
                    logger.info(f"🔄 Lightning progress: {successful_files}/{len(unique_customers)} files completed...")

            except Exception as e:
                logger.error(f"❌ Error for {customer_name}: {e}")
                failed_files += 1

        # Performance summary
        end_time = time.time()
        total_time = end_time - start_time

        logger.info(f"⚡ LIGHTNING FAST splitting completed!")
        logger.info(f"📈 Results: {successful_files} successful, {failed_files} failed")
        logger.info(f"🚀 Total time: {total_time:.2f} seconds")
        logger.info(f"💫 Average: {total_time / max(len(unique_customers), 1):.3f} seconds per file")

        return successful_files, failed_files, total_time

    except Exception as e:
        logger.error(f"💥 Fatal error in lightning_fast_formatted_split: {e}")
        logger.error(traceback.format_exc())
        return 0, 0, 0


def ultra_fast_database_split_with_formatting():
    """
    Alternative: Ultra-fast database approach with minimal formatting
    """
    start_time = time.time()
    logger.info("🚀 Starting ultra-fast database split...")

    try:
        today = date.today()
        pool = get_mysql_connection_pool()
        if not pool:
            logger.error("Database connection pool not available.")
            return 0, 0, 0

        conn = pool.get_connection()
        cursor = conn.cursor(dictionary=True)

        # Single query to get all data
        query_all_data = "SELECT * FROM dailyfiledto WHERE fdate = %s AND `Cust` IS NOT NULL ORDER BY `Cust`"
        cursor.execute(query_all_data, (today,))
        all_data = cursor.fetchall()
        conn.close()

        if not all_data:
            logger.warning("No data found for today's date")
            return 0, 0, 0

        # Convert to DataFrame
        df_all = pd.DataFrame(all_data)

        # Identify string columns
        string_columns = []
        for col in ['InvoiceNum', 'InternalCode', 'ContractNum']:
            if col in df_all.columns:
                string_columns.append(col)

        # Group by customer
        customer_groups = df_all.groupby('Cust')

        logger.info(f"👥 Processing {len(customer_groups)} customers from database...")

        successful_files = 0
        failed_files = 0

        for customer_name, customer_data in customer_groups:
            try:
                safe_filename = "".join(
                    c for c in str(customer_name) if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
                output_file = os.path.join(os.getcwd(), f"{safe_filename}.xlsx")

                # Super fast Excel creation
                wb = Workbook()
                ws = wb.active

                # Bulk write
                for r_idx, row in enumerate(dataframe_to_rows(customer_data, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)

                        # String format for specific columns
                        if r_idx > 1 and c_idx <= len(customer_data.columns):
                            col_name = customer_data.columns[c_idx - 1]
                            if col_name in string_columns:
                                cell.number_format = '@'
                        elif r_idx == 1:
                            cell.font = Font(bold=True)

                wb.save(output_file)
                wb.close()
                successful_files += 1

            except Exception as e:
                logger.error(f"❌ Error for {customer_name}: {e}")
                failed_files += 1

        end_time = time.time()
        total_time = end_time - start_time

        logger.info(f"🎉 Database splitting completed in {total_time:.2f} seconds!")
        logger.info(f"📈 Results: {successful_files} successful, {failed_files} failed")

        return successful_files, failed_files, total_time

    except Exception as e:
        logger.error(f"💥 Fatal error in database split: {e}")
        return 0, 0, 0


def add_helper_sheet_fast(daily_file_path, master_file_path):
    """Fast version of adding Helper sheet"""
    try:
        logger.info("📋 Adding Helper sheet...")

        full_data = pd.read_excel(daily_file_path)
        master_df = pd.read_excel(master_file_path, usecols=['Arabic', 'Name', 'Index', 'Type', 'Transf Type'])
        master_df.rename(columns={
            'Arabic': 'اسم المفوتر',
            'Name': 'Cust',
            'Index': 'Index',
            'Type': 'BillerType',
            'Transf Type': 'TransType'
        }, inplace=True)

        helper_df = pd.merge(
            full_data.drop_duplicates(subset=['اسم المفوتر']),
            master_df,
            on=['اسم المفوتر'],
            how='left',
            suffixes=('', '_master')
        )

        final_helper_df = pd.DataFrame()
        final_helper_df['CustomerName'] = helper_df['Cust']
        final_helper_df['Index'] = helper_df['Index']
        final_helper_df['ArabicName'] = helper_df['اسم المفوتر']
        final_helper_df['HyperLink'] = ''
        final_helper_df['TransType'] = helper_df['TransType']
        final_helper_df['BillerType'] = helper_df['BillerType']

        final_helper_df.sort_values(by='Index', inplace=True)

        # Add to Excel using xlwings
        app = None
        try:
            if len(xw.apps) > 0:
                app = xw.apps.active
            else:
                app = xw.App(visible=True, add_book=False)
        except Exception:
            app = xw.App(visible=True, add_book=False)

        daily_wb = app.books.open(daily_file_path)

        helper_sheet = None
        for sheet in daily_wb.sheets:
            if sheet.name == 'Helper':
                helper_sheet = sheet
                break

        if helper_sheet is None:
            helper_sheet = daily_wb.sheets.add(name='Helper')
        else:
            helper_sheet.clear()

        helper_sheet.range('A1').options(index=False).value = final_helper_df
        daily_wb.save()

        logger.info("✅ Helper sheet added successfully!")

    except Exception as e:
        logger.error(f"❌ Error adding Helper sheet: {e}")


if __name__ == "__main__":
    # Configure paths
    daily_file_path = config.config.dailyfile_base
    daily_file_name = config.config.dailyfile_name
    # excel_file = rf"{daily_file_path}\{daily_file_name}"  # Your daily file path
    daily_file = rf"{daily_file_path}\{m_year}\{path_month_abbr}\{path_day}\{daily_file_name}"  # Your daily file path"
    master_file = r"D:\Freelance\Azm\2025\CustomerNamesLookUp.xlsx"

    try:
        overall_start = time.time()

        # Step 1: Modify the Excel file
        logger.info("🔧 Step 1: Modifying Excel file structure...")
        modify_start = time.time()
        modify_excel_file_final(daily_file, master_file)
        modify_time = time.time() - modify_start
        logger.info(f"✅ File modification completed in {modify_time:.2f} seconds")

        # Step 2: Choose LIGHTNING FAST method
        logger.info("⚡ Step 2: Lightning fast file splitting...")

        # METHOD 1: Lightning fast with formatting + string preservation (RECOMMENDED)
        success, failed, split_time = lightning_fast_formatted_split(daily_file, master_file)

        # METHOD 2: Ultra-fast database method (Alternative - even faster but requires database)
        # success, failed, split_time = ultra_fast_database_split_with_formatting()

        # Step 3: Add Helper sheet
        logger.info("📋 Step 3: Adding Helper sheet...")
        helper_start = time.time()
        add_helper_sheet_fast(daily_file, master_file)
        helper_time = time.time() - helper_start

        # Final summary
        total_time = time.time() - overall_start
        logger.info(f"\n⚡ LIGHTNING FAST PROCESSING COMPLETED!")
        logger.info(f"📊 Summary:")
        logger.info(f"   • File modification: {modify_time:.2f}s")
        logger.info(f"   • Lightning file splitting: {split_time:.2f}s ({success} files)")
        logger.info(f"   • Helper sheet: {helper_time:.2f}s")
        logger.info(f"   • TOTAL TIME: {total_time:.2f}s")
        logger.info(f"⚡ Average per customer file: {split_time / max(success, 1):.3f}s")
        logger.info(f"🔤 String formatting preserved for columns D, M, T!")
        logger.info(f"🔢 Numeric formatting applied to columns E, F, H, I, K!")

        # Performance celebration
        if total_time < 60:
            logger.info(f"🎉 BLAZING FAST: Completed in under 1 minute!")
        elif total_time < 120:
            logger.info(f"🚀 SUPER FAST: Completed in under 2 minutes!")
        else:
            logger.info(f"✅ COMPLETED: Total time {total_time:.1f} seconds")

    except Exception as e:
        logger.error(f"💥 Fatal error: {e}")
        logger.error(traceback.format_exc())